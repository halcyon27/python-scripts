import sys, codecs
from openpyxl import load_workbook

# reads content of template_av.txt, which should be added by user in .\
# template needs placeholders QUERY, INTENT, TAG in proper locations
# concepts must be added manually
def reader(tag, query, intent, tc):
    f = open("template_av.txt", 'r')
    content=f.read()
    content=content.replace('QUERY', query)
    content=content.replace('INTENT', intent)
    content=content.replace('TAG', tag)
    writer(content, tc)
    f.close()

# writes to file named using tc variable
def writer(content, tc):
    fname = str(tc) + '.txt.'
    f = codecs.open(fname, 'w', 'utf-8')
    f.write(content)
    f.close

# loops through excel and reads, writes data by row
def get_xls_data(wb_name, tag):
    wb = load_workbook(wb_name)
    ws = wb.active
    cols = ws.iter_cols(min_row=1, max_row=1,)
    for row in cols:
        for cell in row:
            for n in range(1, ws.max_row):
                # search string must match first row of test case column
                if cell.value == 'TC#':
                    tc = cell.offset(n, 0).value
                    query = cell.offset(n, 4).value
                    intent = cell.offset(n, 5).value
                    reader(tag, query, intent, tc)

def main():
    # argument 1 is name of xlsx
    wb_name = sys.argv[1]
    # argument 2 is nlu context tag for all cases
    tag = sys.argv[2]
    get_xls_data(wb_name, tag)

main()
